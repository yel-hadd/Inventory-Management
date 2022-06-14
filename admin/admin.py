from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.image import Image
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.spinner import Spinner
from kivy.uix.modalview import ModalView
from kivy.uix.boxlayout import BoxLayout
from kivy.lang import Builder

from pymongo import MongoClient
from utils.datatable import DataTable
import hashlib
import openpyxl
import tkinter.filedialog
from datetime import datetime
from  random import choice
from string import digits, ascii_uppercase
from pandas import read_excel, to_numeric
from PIL import ImageFont, ImageDraw
import PIL.Image
import qrcode
import os
import re


Builder.load_file('./admin/admin.kv')


# TIDY UP CODE BY SCREEN
class AdminWindow(BoxLayout):
    # Add screen size to get_products
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        client = MongoClient()
        db = client.pos
        self.users = db.users
        self.products = db.stocks
        self.charges = db.charges
        self.transactions = db.transactions
        self.clients = db.clients
        
        # display users
        content = self.ids.scrn_contents
        users = self.get_users()
        usertable  = DataTable(table=users)
        content.add_widget(usertable)
        
        # display products
        product_scrn = self.ids.scrn_product_contents
        products = self.get_products()
        prod_table  = DataTable(table=products)
        product_scrn.add_widget(prod_table)
        
        #scrn_order_content
        order_scrn = self.ids.scrn_order_contents
        products = self.get_products()
        prod_table  = DataTable(table=products)
        order_scrn.add_widget(prod_table)
        
        #scrn_search_content
        search_scrn = self.ids.scrn_search_contents
        products = self.get_products()
        prod_table  = DataTable(table=products)
        search_scrn.add_widget(prod_table)
        
        # scrn_supplier_contents
        supplier_scrn = self.ids.scrn_supplier_contents
        products = self.get_products()
        prod_table  = DataTable(table=products)
        supplier_scrn.add_widget(prod_table)
        
        # scrn_charges_contents
        charges_scrn = self.ids.scrn_charges_contents
        products = self.get_charges()
        prod_table  = DataTable(table=products)
        charges_scrn.add_widget(prod_table)
        
        # scrn_transaction_contents
        tr_scrn = self.ids.scrn_transaction_contents
        products = self.get_transactions()
        prod_table  = DataTable(table=products)
        tr_scrn.add_widget(prod_table)
        
        # scrn_clients_contents
        tr_scrn = self.ids.scrn_clients_contents
        products = self.get_clients()
        prod_table  = DataTable(table=products)
        tr_scrn.add_widget(prod_table)
        

    def logout(self):
        self.parent.parent.current = "scrn_si"


    def calculate_price(self):
        purchase = self.ids.purchase_price_calc.text
        transport = self.ids.transport_calc.text
        douane = self.ids.douane_calc.text
        transitaire = self.ids.transitaire_calc.text
        autres = self.ids.autres_calc.text
        nbr = self.ids.nbr_prdct_calc.text
        if purchase == '':
            self.error_popup('le champ "Prix D\'achat" manquant')
            return 0
        if nbr == '':
            self.error_popup('le champ "Nombre des Produits" manquant')
            return 0
        if transport == '':
            self.ids.transport_calc.text = "0"
            transport = 0
        if douane == '':
            self.ids.douane_calc.text = "0"
            douane = 0
        if transitaire == '':
            self.ids.transitaire_calc.text = "0"
            transitaire = 0
        if autres == '':
            self.ids.autres_calc.text = "0"
            autres = 0
        try:
            somme = float(purchase)+float(transport)+float(douane)+float(transitaire)+float(autres)
        except ValueError:
            self.error_popup("Entrée invalide")
            return 0
        result = somme/int(nbr)
        result = "{:.2f}".format(result)
        self.ids.result_calc.text = f"Résultat: {result} DH PAR PRODUIT"
        
        return 0
    
    
    def clear_calc(self):
        self.ids.purchase_price_calc.text = ''
        self.ids.transport_calc.text = ''
        self.ids.douane_calc.text = ''
        self.ids.transitaire_calc.text = ''
        self.ids.autres_calc.text = ''
        self.ids.nbr_prdct_calc.text = ''
        self.ids.result_calc.text = "Résultat: "
        
        return 0
    
             
    def create_label(self, ref):
        path = tkinter.filedialog.askdirectory()
        product = self.products.find({"Ref": f"{ref}"})
        try:
            product = product[0]
        except IndexError:
            self.error_popup("Référence Invalide")
            return 0
        prix = product['prix']
        marque = product['marque']
        modele = product['modele']
        cpu = product['cpu']
        ram = product['ram']
        stockage = product['stockage']
        gpu = product['gpu']
        batterie = product['batterie']
        commande = product['commande']

        font = ImageFont.truetype("Montserrat-Regular.ttf", 35)
        font2 = ImageFont.truetype("Montserrat-Regular.ttf", 54)

        img = PIL.Image.open("./utils/Label.jpg")

        draw = ImageDraw.Draw(img)

        draw.text((185, 547),f"{ref}",(0,0,0),font=font2)

        draw.text((397, 1150),f"{marque} {modele}",(0,0,0),font=font)
        draw.text((397, 1224),f"{cpu}",(0,0,0),font=font)
        draw.text((397, 1298),f"{ram}",(0,0,0),font=font)
        draw.text((397, 1372),f"{stockage}",(0,0,0),font=font)
        draw.text((397, 1446),f"{gpu}",(0,0,0),font=font)
        draw.text((397, 1520),f"{batterie}",(0,0,0),font=font)
        draw.text((397, 1594),f"{commande}",(0,0,0),font=font)

        Logo_link = './utils/logo.jpg'

        logo = PIL.Image.open(Logo_link)    

        basewidth = 100 

        wpercent = (basewidth/float(logo.size[0]))
        hsize = int((float(logo.size[1])*float(wpercent)))
        logo = logo.resize((basewidth, hsize), PIL.Image.Resampling.LANCZOS)
        QRcode = qrcode.QRCode( 
	    error_correction=qrcode.constants.ERROR_CORRECT_H
        )
        
        QRcode.add_data(ref)    
        QRcode.make()
        QRcolor = '#000000'
        QRimg = QRcode.make_image(
        	fill_color=QRcolor, back_color="white").convert('RGB')

        QRimg = QRimg.resize((420, 420), PIL.Image.Resampling.NEAREST)
        pos = ((QRimg.size[0] - logo.size[0]) // 2,
        	(QRimg.size[1] - logo.size[1]) // 2)
        QRimg.paste(logo, pos)        
        img.paste(QRimg, (390, 680))

        img.save(f"{path}/{ref}.pdf")
        os.startfile(f"{path}/{ref}.pdf")
        return 0


    def create_order_labels(self, order):
        if order == '':
            return 0
        comm = self.products.find({"commande": f"{order}"})
        try:
            comm = comm[0]
            del(comm)
        except IndexError:
            return 0

        path = tkinter.filedialog.askdirectory()
        if path == '':
            return 0

        font = ImageFont.truetype("Montserrat-Regular.ttf", 35)
        font2 = ImageFont.truetype("Montserrat-Regular.ttf", 54)
        img = PIL.Image.open("./utils/Label.jpg")
        Logo_link = './utils/logo.jpg'
        images = []
        for product in self.products.find({"commande": f"{order}"}):
            ref = product['Ref']
            prix = product['prix']
            marque = product['marque']
            modele = product['modele']
            cpu = product['cpu']
            ram = product['ram']
            stockage = product['stockage']
            gpu = product['gpu']
            batterie = product['batterie']
            commande = product['commande']

            draw = ImageDraw.Draw(img)
            draw.text((185, 547),f"{ref}",(0,0,0),font=font2)
            draw.text((397, 1150),f"{marque} {modele}",(0,0,0),font=font)
            draw.text((397, 1224),f"{cpu}",(0,0,0),font=font)
            draw.text((397, 1298),f"{ram}",(0,0,0),font=font)
            draw.text((397, 1372),f"{stockage}",(0,0,0),font=font)
            draw.text((397, 1446),f"{gpu}",(0,0,0),font=font)
            draw.text((397, 1520),f"{batterie}",(0,0,0),font=font)
            draw.text((397, 1594),f"{commande}",(0,0,0),font=font)

            logo = PIL.Image.open(Logo_link)    
            basewidth = 100 
            wpercent = (basewidth/float(logo.size[0]))
            hsize = int((float(logo.size[1])*float(wpercent)))
            logo = logo.resize((basewidth, hsize), PIL.Image.Resampling.LANCZOS)
            QRcode = qrcode.QRCode( 
            error_correction=qrcode.constants.ERROR_CORRECT_H
            )   
            QRcode.add_data(ref)    
            QRcode.make()
            QRcolor = '#000000'
            QRimg = QRcode.make_image(
                fill_color=QRcolor, back_color="white").convert('RGB')
            QRimg = QRimg.resize((420, 420), PIL.Image.Resampling.NEAREST)
            pos = ((QRimg.size[0] - logo.size[0]) // 2,
                (QRimg.size[1] - logo.size[1]) // 2)
            QRimg.paste(logo, pos)        
            img.paste(QRimg, (390, 680))
            images.append(img)
            img = PIL.Image.open("./utils/Label.jpg")

        images[0].save(
            f"{path}/{order}.pdf", "PDF" ,resolution=300.0, save_all=True, append_images=images[1:]
        )
        
        os.startfile(f"{path}/{order}.pdf")
        self.success_popup("Tickets Générées")
        return 0


    def get_order(self, for_export=False):
        orderref = self.ids.order_id.text
        if orderref == '':
            return 0
        products = self.products

        if for_export:
            _stocks = {}
            _stocks['Ref'] = {}
            _stocks['marque'] = {}
            _stocks['modele'] = {}
            _stocks['cpu'] = {}
            _stocks['ram'] = {}
            _stocks['stockage'] = {}
            _stocks['gpu'] = {}
            _stocks['batterie'] = {}
            
            _stocks["cout d'achat"] = {}
            _stocks['prix de vente'] = {}
            _stocks['gain'] = {}
            _stocks['en_stock'] = {}
            _stocks['vendu'] = {}
            _stocks['commande'] = {}
            _stocks['fournisseur'] = {}
            _stocks['dernier_achat'] = {}
            _stocks['commentaire'] = {}

            Ref = []
            prix = []
            prix_achat = []
            gain = []
            marque = []
            modele = []
            cpu = []
            ram = []
            gpu = []
            stockage = []
            batterie = []
            en_stock = []
            vendu = []
            commande = []
            fournisseur = []
            dernier_achat = []
            commentaire = []

            for product in products.find({"commande": f"{orderref}"}):
                Ref.append(product['Ref'])
                try:
                    prix.append(product['prix'])
                except:
                    prix.append('0')
                try:
                    prix_achat.append(product['prix_achat'])
                except KeyError:
                    prix_achat.append('0')

                try:
                    tbuy_p = float(product['prix_achat'])
                except:
                    tbuy_p = 0
                    
                try:
                    tsell_p = float(product['prix'])
                except:
                    tsell_p = 0
                
                gain.append(tsell_p-tbuy_p)
                
                marque.append(product['marque'])
                modele.append(product['modele'])
                cpu.append(product['cpu'])
                ram.append(product['ram'])
                gpu.append(product['gpu'])
                stockage.append(product['stockage'])
                batterie.append(product['batterie'])
                try:    
                    en_stock.append(product['en_stock'])
                except KeyError:
                    en_stock.append('')

                try:    
                    vendu.append(product['vendu'])
                except KeyError:
                    vendu.append('')

                commande.append(product['commande'])

                try:    
                    fournisseur.append(product['fournisseur'])
                except KeyError:
                    fournisseur.append('')

                try:    
                    dernier_achat.append(product['dernier_achat'])
                except KeyError:
                    dernier_achat.append('')

                try:    
                    commentaire.append(product['commentaire'])
                except KeyError:
                    commentaire.append('****')        

            sum_purchase_price = []
            sum_sold_list = []
            sold_items = []
            in_stock_items = []
            in_stock_items_price = []
            
            for c, v in enumerate(Ref):
                _stocks['Ref'][c] = Ref[c]
                _stocks['marque'][c] = marque[c]
                _stocks['modele'][c] = modele[c]
                _stocks['cpu'][c] = cpu[c]
                _stocks['ram'][c] = ram[c]
                _stocks['stockage'][c] = stockage[c]
                _stocks['gpu'][c] = gpu[c]
                _stocks['batterie'][c] = batterie[c]
                _stocks['prix de vente'][c] = prix[c]
                _stocks["cout d'achat"][c] = prix_achat[c]
                _stocks["gain"][c] = gain[c]
                _stocks['en_stock'][c] = en_stock[c]
                _stocks['vendu'][c] = vendu[c]
                _stocks['commande'][c] = commande[c]
                _stocks['dernier_achat'][c] = dernier_achat[c]
                _stocks['commentaire'][c] = commentaire[c]
                _stocks['fournisseur'][c] = fournisseur[c]
            
                if vendu[c] == '':
                    vendu[c] = 0

                if en_stock[c] == '':
                    en_stock[c] = 0
        else:
            stats = self.ids.scrn_order_stats
            order_scrn = self.ids.scrn_order_contents
            order_scrn.clear_widgets()
            stats.clear_widgets()
            _stocks = {}
            _stocks['Ref'] = {}
            _stocks['designation'] = {}
            _stocks["cout d'achat"] = {}
            _stocks['prix de vente'] = {}
            _stocks['gain'] = {}
            _stocks['en_stock'] = {}
            _stocks['vendu'] = {}
            _stocks['commande'] = {}
            _stocks['fournisseur'] = {}
            _stocks['dernier_achat'] = {}
            _stocks['commentaire'] = {}

            Ref = []
            prix = []
            prix_achat = []
            marque = []
            modele = []
            cpu = []
            ram = []
            gpu = []
            stockage = []
            batterie = []
            en_stock = []
            vendu = []
            commande = []
            fournisseur = []
            dernier_achat = []
            commentaire = []
            gain = []
            charges = []

            for product in products.find({"commande": f"{orderref}"}):
                Ref.append(product['Ref'])
                prix.append(product['prix'])
                try:
                    prix_achat.append(product['prix_achat'])
                except KeyError:
                    prix_achat.append('')
                marque.append(product['marque'])
                modele.append(product['modele'])
                cpu.append(product['cpu'])
                ram.append(product['ram'])
                gpu.append(product['gpu'])
                stockage.append(product['stockage'])
                batterie.append(product['batterie'])
                try:    
                    en_stock.append(product['en_stock'])
                except KeyError:
                    en_stock.append('')

                try:    
                    vendu.append(product['vendu'])
                except KeyError:
                    vendu.append('')

                commande.append(product['commande'])

                try:    
                    fournisseur.append(product['fournisseur'])
                except KeyError:
                    fournisseur.append('')

                try:    
                    dernier_achat.append(product['dernier_achat'])
                except KeyError:
                    dernier_achat.append('')

                try:    
                    commentaire.append(product['commentaire'])
                except KeyError:
                    commentaire.append('****')
                
                try:
                    tbuy_p = float(product['prix_achat'])
                except:
                    tbuy_p = 0
                    
                try:
                    tsell_p = float(product['prix'])
                except:
                    tsell_p = 0
                
                gain.append(tsell_p-tbuy_p)
                
                try:
                    charges.append(float(product['charges']))
                except:
                    charges.append(0)
                    
            sum_purchase_price = []
            sum_sold_list = []
            sold_items = []
            in_stock_items = []
            in_stock_items_price = []

            for c, v in enumerate(Ref):
                _stocks['Ref'][c] = Ref[c]
                _stocks['designation'][c] = f"{marque[c]} {modele[c]} | {cpu[c]} | {ram[c]}GB\n{stockage[c]} | {gpu[c]} | {batterie[c]}"
                _stocks['prix de vente'][c] = prix[c]
                _stocks["cout d'achat"][c] = prix_achat[c]
                _stocks['en_stock'][c] = en_stock[c]
                _stocks['vendu'][c] = vendu[c]
                _stocks['commande'][c] = commande[c]
                _stocks['dernier_achat'][c] = dernier_achat[c]
                _stocks['commentaire'][c] = commentaire[c]
                _stocks['fournisseur'][c] = fournisseur[c]
                _stocks['gain'][c] = gain[c]
            
                if vendu[c] == '':
                    vendu[c] = 0

                if en_stock[c] == '':
                    en_stock[c] = 0

                if int(vendu[c]) >= 1:
                    sum_sold_list.append((int(vendu[c])*float(prix[c]))-charges[c])
                    sold_items.append(int(vendu[c]))
                    sum_purchase_price.append((int(vendu[c])*float(prix_achat[c]))-charges[c])

                if int(en_stock[c]) >= 1:
                    in_stock_items_price.append((int(en_stock[c])*float(prix_achat[c]))-charges[c])
                    in_stock_items.append(int(en_stock[c]))
                    sum_purchase_price.append((int(en_stock[c])*float(prix_achat[c]))-charges[c])

            products = _stocks
            prod_table  = DataTable(table=products)
            order_scrn.add_widget(prod_table)
            
            # key numbers
            total_order_price = sum(sum_purchase_price)
            in_stock_items = sum(in_stock_items)
            in_stock_items_price = sum(in_stock_items_price)
            sold_items = sum(sold_items)
            sold_products_price = float(sum(sum_sold_list))
            profit = sold_products_price-total_order_price-sum(charges)

            totalorderpricelabel = Label(text=f'Prix Totale:\n{total_order_price} DH', bold=True, color=(0,0,0,1))
            instockitemslabel = Label(text=f'Produits en stock:\n{in_stock_items}', bold=True, color=(0,0,0,1))
            instockitemspricelabel = Label(text=f'Prix des produits en stock:\n{in_stock_items_price} DH', bold=True, color=(0,0,0,1))
            solditemslabel = Label(text=f'Produits Vendu:\n{sold_items}', bold=True, color=(0,0,0,1))
            soldproductspricelabel = Label(text=f'Prix des produits Vendu:\n{sold_products_price} DH', bold=True, color=(0,0,0,1))
            profitlabel = Label(text=f'Gain:\n{profit} DH', bold=True, color=(0,0,0,1))
            chargeslabel = Label(text=f'Frais:\n{sum(charges)} DH', bold=True, color=(0,0,0,1))

            stats.add_widget(totalorderpricelabel)
            stats.add_widget(chargeslabel)
            stats.add_widget(profitlabel)
            stats.add_widget(instockitemslabel)
            stats.add_widget(instockitemspricelabel)
            stats.add_widget(solditemslabel)
            stats.add_widget(soldproductspricelabel)
            
        return _stocks


    def get_supp(self, for_export=False):
        orderref = self.ids.supplier_id.text
        if orderref == '':
            return 0
        products = self.products

        
        if for_export:
            _stocks = {}
            _stocks['Ref'] = {}
            _stocks['marque'] = {}
            _stocks['modele'] = {}
            _stocks['cpu'] = {}
            _stocks['ram'] = {}
            _stocks['stockage'] = {}
            _stocks['gpu'] = {}
            _stocks['batterie'] = {}
            _stocks["cout d'achat"] = {}
            _stocks['prix de vente'] = {}
            _stocks['gain'] = {}
            _stocks['en_stock'] = {}
            _stocks['vendu'] = {}
            _stocks['commande'] = {}
            _stocks['fournisseur'] = {}
            _stocks['dernier_achat'] = {}
            _stocks['commentaire'] = {}

            Ref = []
            prix = []
            prix_achat = []
            marque = []
            modele = []
            cpu = []
            ram = []
            gpu = []
            stockage = []
            batterie = []
            en_stock = []
            vendu = []
            commande = []
            fournisseur = []
            dernier_achat = []
            commentaire = []
            gain = []

            for product in products.find({"fournisseur": f"{orderref}"}):
                Ref.append(product['Ref'])
                prix.append(product['prix'])
                try:
                    prix_achat.append(product['prix_achat'])
                except KeyError:
                    prix_achat.append('')
                marque.append(product['marque'])
                modele.append(product['modele'])
                cpu.append(product['cpu'])
                ram.append(product['ram'])
                gpu.append(product['gpu'])
                stockage.append(product['stockage'])
                batterie.append(product['batterie'])
                try:    
                    en_stock.append(product['en_stock'])
                except KeyError:
                    en_stock.append('')

                try:    
                    vendu.append(product['vendu'])
                except KeyError:
                    vendu.append('')

                commande.append(product['commande'])

                try:    
                    fournisseur.append(product['fournisseur'])
                except KeyError:
                    fournisseur.append('')

                try:    
                    dernier_achat.append(product['dernier_achat'])
                except KeyError:
                    dernier_achat.append('')

                try:    
                    commentaire.append(product['commentaire'])
                except KeyError:
                    commentaire.append('****')
                
                try:
                    tbuy_p = float(product['prix_achat'])
                except:
                    tbuy_p = 0

                try:
                    tsell_p = float(product['prix'])
                except:
                    tsell_p = 0

                gain.append(tsell_p-tbuy_p)

            sum_purchase_price = []
            sum_sold_list = []
            sold_items = []
            in_stock_items = []
            in_stock_items_price = []

            for c, v in enumerate(Ref):
                _stocks['Ref'][c] = Ref[c]
                _stocks['marque'][c] = marque[c]
                _stocks['modele'][c] = modele[c]
                _stocks['cpu'][c] = cpu[c]
                _stocks['ram'][c] = ram[c]
                _stocks['stockage'][c] = stockage[c]
                _stocks['gpu'][c] = gpu[c]
                _stocks['batterie'][c] = batterie[c]
                _stocks['prix de vente'][c] = prix[c]
                _stocks["cout d'achat"][c] = prix_achat[c]
                _stocks['en_stock'][c] = en_stock[c]
                _stocks['vendu'][c] = vendu[c]
                _stocks['commande'][c] = commande[c]
                _stocks['dernier_achat'][c] = dernier_achat[c]
                _stocks['commentaire'][c] = commentaire[c]
                _stocks['fournisseur'][c] = fournisseur[c]
                _stocks['gain'][c] = gain[c]
                
            
                if vendu[c] == '':
                    vendu[c] = 0

                if en_stock[c] == '':
                    en_stock[c] = 0
        else:
            stats = self.ids.scrn_supplier_stats
            order_scrn = self.ids.scrn_supplier_contents
            order_scrn.clear_widgets()
            stats.clear_widgets()
            _stocks = {}
            _stocks['Ref'] = {}
            _stocks['designation'] = {}
            _stocks["cout d'achat"] = {}
            _stocks['prix de vente'] = {}
            _stocks['gain'] = {}
            _stocks['en_stock'] = {}
            _stocks['vendu'] = {}
            _stocks['commande'] = {}
            _stocks['fournisseur'] = {}
            _stocks['dernier_achat'] = {}
            _stocks['commentaire'] = {}

            Ref = []
            prix = []
            prix_achat = []
            marque = []
            modele = []
            cpu = []
            ram = []
            gpu = []
            stockage = []
            batterie = []
            en_stock = []
            vendu = []
            commande = []
            fournisseur = []
            dernier_achat = []
            commentaire = []
            gain = []

            for product in products.find({"fournisseur": f"{orderref}"}):
                Ref.append(product['Ref'])
                prix.append(product['prix'])
                try:
                    prix_achat.append(product['prix_achat'])
                except KeyError:
                    prix_achat.append('')
                marque.append(product['marque'])
                modele.append(product['modele'])
                cpu.append(product['cpu'])
                ram.append(product['ram'])
                gpu.append(product['gpu'])
                stockage.append(product['stockage'])
                batterie.append(product['batterie'])
                try:    
                    en_stock.append(product['en_stock'])
                except KeyError:
                    en_stock.append('')

                try:    
                    vendu.append(product['vendu'])
                except KeyError:
                    vendu.append('')

                commande.append(product['commande'])

                try:    
                    fournisseur.append(product['fournisseur'])
                except KeyError:
                    fournisseur.append('')

                try:    
                    dernier_achat.append(product['dernier_achat'])
                except KeyError:
                    dernier_achat.append('')

                try:    
                    commentaire.append(product['commentaire'])
                except KeyError:
                    commentaire.append('****')
                try:
                    tbuy_p = float(product['prix_achat'])
                except:
                    tbuy_p = 0

                try:
                    tsell_p = float(product['prix'])
                except:
                    tsell_p = 0

                gain.append(tsell_p-tbuy_p)

            sum_purchase_price = []
            sum_sold_list = []
            sold_items = []
            in_stock_items = []
            in_stock_items_price = []

            for c, v in enumerate(Ref):
                _stocks['Ref'][c] = Ref[c]
                _stocks['designation'][c] = f"{marque[c]} {modele[c]} | {cpu[c]} | {ram[c]}GB\n{stockage[c]} | {gpu[c]} | {batterie[c]}"
                _stocks['prix de vente'][c] = prix[c]
                _stocks["cout d'achat"][c] = prix_achat[c]
                _stocks['en_stock'][c] = en_stock[c]
                _stocks['vendu'][c] = vendu[c]
                _stocks['commande'][c] = commande[c]
                _stocks['dernier_achat'][c] = dernier_achat[c]
                _stocks['commentaire'][c] = commentaire[c]
                _stocks['fournisseur'][c] = fournisseur[c]
                _stocks['gain'][c] = gain[c]
            
                if vendu[c] == '':
                    vendu[c] = 0

                if en_stock[c] == '':
                    en_stock[c] = 0

                if int(vendu[c]) >= 1:
                    sum_sold_list.append(int(vendu[c])*float(prix[c]))
                    sold_items.append(int(vendu[c]))
                    sum_purchase_price.append(int(vendu[c])*float(prix_achat[c]))

                if int(en_stock[c]) >= 1:
                    in_stock_items_price.append(int(en_stock[c])*float(prix_achat[c]))
                    in_stock_items.append(int(en_stock[c]))
                    sum_purchase_price.append(int(en_stock[c])*float(prix_achat[c]))

            products = _stocks
            prod_table  = DataTable(table=products)
            order_scrn.add_widget(prod_table)
            
            # key numbers
            total_order_price = sum(sum_purchase_price)
            in_stock_items = sum(in_stock_items)
            in_stock_items_price = sum(in_stock_items_price)
            sold_items = sum(sold_items)
            sold_products_price = float(sum(sum_sold_list))
            profit = sold_products_price-total_order_price

            totalorderpricelabel = Label(text=f'Prix Totale:\n{total_order_price} DH', bold=True, color=(0,0,0,1))
            instockitemslabel = Label(text=f'Produits en stock:\n{in_stock_items}', bold=True, color=(0,0,0,1))
            instockitemspricelabel = Label(text=f'Prix des produits en stock:\n{in_stock_items_price} DH', bold=True, color=(0,0,0,1))
            solditemslabel = Label(text=f'Produits Vendu:\n{sold_items}', bold=True, color=(0,0,0,1))
            soldproductspricelabel = Label(text=f'Prix des produits Vendu:\n{sold_products_price} DH', bold=True, color=(0,0,0,1))
            profitlabel = Label(text=f'Profit:\n{profit} DH', bold=True, color=(0,0,0,1))

            stats.add_widget(totalorderpricelabel)
            stats.add_widget(instockitemslabel)
            stats.add_widget(instockitemspricelabel)
            stats.add_widget(solditemslabel)
            stats.add_widget(soldproductspricelabel)
            stats.add_widget(profitlabel)
        return _stocks


    def get_product(self, order='False'):
        if order == 'False':
            orderref = self.ids.prod_id.text
        else:
            orderref = order
        
        if orderref == '':
            return 0
        
        if order != 'False':
            _stocks = {}
            _stocks['Ref'] = {}
            _stocks['marque'] = {}
            _stocks['modele'] = {}
            _stocks['cpu'] = {}
            _stocks['ram'] = {}
            _stocks['stockage'] = {}
            _stocks['gpu'] = {}
            _stocks['batterie'] = {} 
            _stocks["cout d'achat"] = {}
            _stocks['prix de vente'] = {}
            _stocks['gain'] = {}
            _stocks['en_stock'] = {}
            _stocks['vendu'] = {}
            _stocks['commande'] = {}
            _stocks['fournisseur'] = {}
            _stocks['dernier_achat'] = {}
            _stocks['commentaire'] = {}
            _stocks['charges'] = {}

            Ref = []
            prix = []
            prix_achat = []
            marque = []
            modele = []
            cpu = []
            ram = []
            gpu = []
            stockage = []
            batterie = []
            en_stock = []
            vendu = []
            commande = []
            fournisseur = []
            dernier_achat = []
            commentaire = []
            gain = []
            charges = []

            for product in self.products.find({"Ref": f"{orderref}"}):
                Ref.append(product['Ref'])
                prix.append(product['prix'])
                try:
                    prix_achat.append(product['prix_achat'])
                except KeyError:
                    prix_achat.append('')
                marque.append(product['marque'])
                modele.append(product['modele'])
                cpu.append(product['cpu'])
                ram.append(product['ram'])
                gpu.append(product['gpu'])
                stockage.append(product['stockage'])
                batterie.append(product['batterie'])
                try:    
                    en_stock.append(product['en_stock'])
                except KeyError:
                    en_stock.append('')

                try:    
                    vendu.append(product['vendu'])
                except KeyError:
                    vendu.append('')

                commande.append(product['commande'])
                try:
                    charges.append(product['charges'])
                except:
                    charges.append(0)

                try:    
                    fournisseur.append(product['fournisseur'])
                except KeyError:
                    fournisseur.append('')

                try:    
                    dernier_achat.append(product['dernier_achat'])
                except KeyError:
                    dernier_achat.append('')

                try:    
                    commentaire.append(product['commentaire'])
                except KeyError:
                    commentaire.append('****')
                    
                try:
                    tbuy_p = float(product['prix_achat'])
                except:
                    tbuy_p = 0
                    
                try:
                    tsell_p = float(product['prix'])
                except:
                    tsell_p = 0
                
                gain.append(tsell_p-tbuy_p)

            for c, v in enumerate(Ref):
                _stocks['Ref'] = Ref[c]
                _stocks['marque'] = marque[c]
                _stocks['modele'] = modele[c]
                _stocks['cpu'] = cpu[c]
                _stocks['ram'] = ram[c]
                _stocks['stockage'] = stockage[c]
                _stocks['gpu'] = gpu[c]
                _stocks['batterie'] = batterie[c]
                _stocks['prix de vente'] = prix[c]
                _stocks["cout d'achat"] = prix_achat[c]
                _stocks['en_stock'] = en_stock[c]
                _stocks['vendu'] = vendu[c]
                _stocks['commande'] = commande[c]
                _stocks['dernier_achat'] = dernier_achat[c]
                _stocks['commentaire'] = commentaire[c]
                _stocks['fournisseur'] = fournisseur[c]
                _stocks['gain'] = gain[c]
                _stocks['charges'] = charges[c]

                if vendu[c] == '':
                    vendu[c] = 0

                if en_stock[c] == '':
                    en_stock[c] = 0
            
            return _stocks
        
        order_scrn = self.ids.scrn_search_contents
        order_scrn.clear_widgets()
        
        _stocks = {}
        _stocks['Ref'] = {}
        _stocks['designation'] = {}
        
        _stocks["cout d'achat"] = {}
        _stocks['prix de vente'] = {}
        _stocks['gain'] = {}
        _stocks['en_stock'] = {}
        _stocks['vendu'] = {}
        _stocks['commande'] = {}
        _stocks['fournisseur'] = {}
        _stocks['dernier_achat'] = {}
        _stocks['commentaire'] = {}
        _stocks['charges'] = {}
        Ref = []
        prix = []
        prix_achat = []
        marque = []
        modele = []
        cpu = []
        ram = []
        gpu = []
        stockage = []
        batterie = []
        en_stock = []
        vendu = []
        commande = []
        fournisseur = []
        dernier_achat = []
        commentaire = []
        gain = []
        charges = []
        
        for product in self.products.find({"Ref": f"{orderref}"}):
            Ref.append(product['Ref'])
            prix.append(product['prix'])

            try:
                prix_achat.append(product['prix_achat'])
            except KeyError:
                prix_achat.append('')

            marque.append(product['marque'])
            modele.append(product['modele'])
            cpu.append(product['cpu'])
            ram.append(product['ram'])
            gpu.append(product['gpu'])
            stockage.append(product['stockage'])
            batterie.append(product['batterie'])

            try:    
                en_stock.append(product['en_stock'])
            except KeyError:
                en_stock.append('')

            try:    
                vendu.append(product['vendu'])
            except KeyError:
                vendu.append('')

            commande.append(product['commande'])
            
            try:    
                fournisseur.append(product['fournisseur'])
            except KeyError:
                fournisseur.append('')

            try:    
                dernier_achat.append(product['dernier_achat'])
            except KeyError:
                dernier_achat.append('')
                
            try:    
                commentaire.append(product['commentaire'])
            except KeyError:
                commentaire.append('')
                
            try:
                tbuy_p = float(product['prix_achat'])
            except:
                tbuy_p = 0
                
            try:
                tsell_p = float(product['prix'])
            except:
                tsell_p = 0
                
            try:
                charges.append(product['charges'])
            except:
                charges.append(0)
                
            gain.append(tsell_p-tbuy_p)
        
        for c, v in enumerate(Ref):
            _stocks['Ref'][c] = Ref[c]
            _stocks['designation'][c] = f"{marque[c]} {modele[c]} | {cpu[c]} | {ram[c]}GB\n{stockage[c]} | {gpu[c]} | {batterie[c]}"
            _stocks['prix de vente'][c] = prix[c]
            _stocks["cout d'achat"][c] = prix_achat[c]
            _stocks['en_stock'][c] = en_stock[c]
            _stocks['vendu'][c] = vendu[c]
            _stocks['commande'][c] = commande[c]
            _stocks['fournisseur'][c] = fournisseur[c]
            _stocks['dernier_achat'][c] = dernier_achat[c]
            _stocks['commentaire'][c] = commentaire[c]
            _stocks['gain'][c] = gain[c]
            _stocks['charges'] = charges[c]
        
        products = _stocks
        prod_table  = DataTable(table=products)
        order_scrn.add_widget(prod_table)
        
        return _stocks

    
    def missing_field_popup(self, field):
            box = BoxLayout(orientation='vertical', spacing=10, size_hint_y=None)        
            error_img = Image(source='./utils/1.png', size=(150, 150))
            error_msg = Label(text=f'champ obligatoire ({field}) manquant', bold=True)
            
            
            popup = ModalView(size_hint=(None, None), size=(400, 300))
            box.add_widget(error_img)
            box.add_widget(error_msg)
            
            popup.add_widget(box)
            
            popup.open()
            
            return 0
    
    
    def error_popup(self, message):
            box = BoxLayout(orientation='vertical', spacing=10, size_hint_y=None)        
            error_img = Image(source='./utils/1.png', size=(150, 150))
            error_msg = Label(text=f'{message}', bold=True)
            
            popup = ModalView(size_hint=(None, None), size=(400, 300))
            box.add_widget(error_img)
            box.add_widget(error_msg)
            
            popup.add_widget(box)
            
            popup.open()
            
            return 0
    
    
    def success_popup(self, operation):
        box = BoxLayout(orientation='vertical', spacing=10, size_hint_y=None)        
        error_img = Image(source='./utils/2.png', size=(150, 150))
        error_msg = Label(text=f'{operation} avec succès', bold=True)
        
        
        popup = ModalView(size_hint=(None, None), size=(400, 300))
        box.add_widget(error_img)
        box.add_widget(error_msg)
        
        popup.add_widget(box)
        
        popup.open()
        
        return 0


    def add_print_fields(self):
        target = self.ids.ops_fields_p
        target.clear_widgets()
        crud_ref = TextInput(hint_text="Référence", multiline=False, width=100, height=10)
        crud_submit =  Button(text='Imprimer Le Ticket', size_hint_x=None, width=150, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.create_label(crud_ref.text))
        crud_close =  Button(text='Fermer', size_hint_x=None, width=100, on_release=lambda x: self.ids.ops_fields_p.clear_widgets(),
                            background_color=(0.184, 0.216, 0.231), background_normal='')
        spacer = Label(text='', size_hint_x=.6)
        target.add_widget(crud_ref)
        target.add_widget(crud_submit)
        target.add_widget(crud_close)
        target.add_widget(spacer)
        

    def gen_ref(self, marque):
        numbers = digits
        letters = ascii_uppercase

        g = True

        while g == True:
            one = ''.join(choice(numbers) for i in range(4))
            two = ''.join(choice(letters) for i in range(1))
            ref = f"{marque[:3]}-{one}{two}"
            g = self.product_exist(ref)
        
        return ref
    
    
    def gen_charge_ref(self):
        numbers = digits
        letters = ascii_uppercase

        g = True

        while g == True:
            one = ''.join(choice(numbers) for i in range(5))
            two = ''.join(choice(letters) for i in range(1))
            ref = f"CH-{one}{two}"
            g = self.charge_exist(ref)
        
        return ref
   

    def export_order_xlsx(self, enstock=False):
        orderref = self.ids.order_id.text
        if orderref == '':
            return 0
        
        path = tkinter.filedialog.askdirectory()
        if path == '':
            return 0
        
        
        _stocks = self.get_order(for_export=True)

        titles = _stocks.keys()
        titles = list(titles)

        titles2 = ['Ref', 'Marque', 'Modele', 'CPU', 'RAM', 'Stockage', 'GPU', 'Batterie', 'Prix De Vente',
                   "cout d'achat", 'En stock', 'Vendu', 'Commande', 'Fournisseur', 'Dernier Achat',
                   'Commentaire']

        product = []
        all_products = []

        for c, v in enumerate(_stocks['Ref']):
            for count, value in enumerate(titles):
                product.append(_stocks[titles[count]][c])
            all_products.append(product)
            product = []

        if enstock == True:
            filename = f'/en-stock-{orderref}-'
            filtered = []
            for i in all_products:
                try:
                    if int(i[10]) >= 1:
                        filtered.append(i)
                except ValueError:
                    i[10] = 0 
                    if int(i[10]) >= 1:
                        filtered.append(i)
            all_products = filtered
        else: 
            filename = f'/tous-{orderref}-'

        all_products.insert(0, titles2)

        wb = openpyxl.Workbook()
        ws_write = wb.worksheets[0]

        for p in all_products:
            ws_write.append(p)

        today = datetime.today().strftime("%d-%m-%Y")
        wb.save(filename=f'{path}{filename}{today}.xlsx')
        
        if enstock:
            self.success_popup('les produits en stock ont été exportés')
        else:
            self.success_popup('Tous les produits ont été exportés')
        
        return 0


    def export_supp_xlsx(self, enstock=False):
        orderref = self.ids.supplier_id.text
        if orderref == '':
            return 0
        
        path = tkinter.filedialog.askdirectory()
        if path == '':
            return 0
        
        
        _stocks = self.get_supp(for_export=True)

        titles = _stocks.keys()
        titles = list(titles)

        titles2 = ['Ref', 'Marque', 'Modele', 'CPU', 'RAM', 'Stockage', 'GPU', 'Batterie', 'Prix De Vente',
                   "cout d'achat", 'En stock', 'Vendu', 'Commande', 'Fournisseur', 'Dernier Achat',
                   'Commentaire']

        product = []
        all_products = []

        for c, v in enumerate(_stocks['Ref']):
            for count, value in enumerate(titles):
                product.append(_stocks[titles[count]][c])
            all_products.append(product)
            product = []

        if enstock == True:
            filename = f'/en-stock-{orderref}-'
            filtered = []
            for i in all_products:
                try:
                    if int(i[10]) >= 1:
                        filtered.append(i)
                except ValueError:
                    i[10] = 0
            all_products = filtered
        else: 
            filename = f'/tous-{orderref}-'

        all_products.insert(0, titles2)

        wb = openpyxl.Workbook()
        ws_write = wb.worksheets[0]

        for p in all_products:
            ws_write.append(p)

        today = datetime.today().strftime("%d-%m-%Y")
        wb.save(filename=f'{path}{filename}{today}.xlsx')
        
        if enstock:
            self.success_popup('les produits en stock ont été exportés')
        else:
            self.success_popup('Tous les produits ont été exportés')
        
        return 0


    def export_xlsx(self, enstock=False):
        path = tkinter.filedialog.askdirectory()
        if path == '':
            return 0
        
        _stocks = self.get_products(for_export=True)
        
        titles = _stocks.keys()
        titles = list(titles)

        titles2 = ['Ref', 'Marque', 'Modele', 'CPU', 'RAM', 'Stockage', 'GPU', 'Batterie', 'Prix De Vente',
                   "cout d'achat", 'En stock', 'Vendu', 'Commande', 'Fournisseur', 'Dernier Achat',
                   'Commentaire']

        product = []
        all_products = []

        for c, v in enumerate(_stocks['Ref']):
            for count, value in enumerate(titles):
                product.append(_stocks[titles[count]][c])
            all_products.append(product)
            product = []
            
        if enstock == True:
            filename = '/en-stock-'
            filtered = []
            for i in all_products:
                if int(i[10]) >= 1:
                    filtered.append(i)
            all_products = filtered
        else:
            filename = '/tous-'

        all_products.insert(0, titles2)

        wb = openpyxl.Workbook()
        ws_write = wb.worksheets[0]

        for p in all_products:
            ws_write.append(p)

        today = datetime.today().strftime("%d-%m-%Y")
        wb.save(filename=f'{path}{filename}{today}.xlsx')
        
        if enstock:
            self.success_popup('les produits en stock ont été exportés')
        else:
            self.success_popup('Tous les produits ont été exportés')
        
        return 0


    def export_charge_xlsx(self):
        path = tkinter.filedialog.askdirectory()
        if path == '':
            return 0
        
        _stocks = self.get_charges()
        
        titles = _stocks.keys()
        titles = list(titles)

        titles2 = ['Ref', 'Montant', 'Motif', 'Date']

        product = []
        all_products = []

        for c, v in enumerate(_stocks['Ref']):
            for count, value in enumerate(titles):
                product.append(_stocks[titles[count]][c])
            all_products.append(product)
            product = []
            
        
        filename = '/charges-'

        all_products.insert(0, titles2)

        wb = openpyxl.Workbook()
        ws_write = wb.worksheets[0]

        for p in all_products:
            ws_write.append(p)

        today = datetime.today().strftime("%d-%m-%Y")
        wb.save(filename=f'{path}{filename}{today}.xlsx')
        
        self.success_popup('Tous les charges ont été exportés')
        
        return 0


    def export_transactions_xlsx(self):
        path = tkinter.filedialog.askdirectory()
        if path == '':
            return 0
        
        _stocks = self.get_transactions()
        
        titles = _stocks.keys()
        titles = list(titles)

        titles2 = ['Ref', 'Totale HT', 'TVA', 'Totale TTC', 'Date']

        product = []
        all_products = []

        for c, v in enumerate(_stocks['Ref']):
            for count, value in enumerate(titles):
                product.append(_stocks[titles[count]][c])
            all_products.append(product)
            product = []
            
        
        filename = '/transactions-'

        all_products.insert(0, titles2)

        wb = openpyxl.Workbook()
        ws_write = wb.worksheets[0]

        for p in all_products:
            ws_write.append(p)

        today = datetime.today().strftime("%d-%m-%Y")
        wb.save(filename=f'{path}{filename}{today}.xlsx')
        
        self.success_popup('Tous les transactions ont été exportés')
        
        return 0


    def import_xlsx(self):
        path = tkinter.filedialog.askopenfilename(initialdir="/", title="Sélectionnez un fichier Excel",
                                                filetypes=[("Fichiers Excel", ".xlsx")])
        if path == '':
            return 0
        
        
        content = self.ids.scrn_product_contents
        content.clear_widgets()
        
        order_scrn = self.ids.scrn_order_contents
        order_scrn.clear_widgets()
        
        
        supplier_scrn = self.ids.scrn_supplier_contents
        supplier_scrn.clear_widgets()
        
        search_scrn = self.ids.scrn_search_contents
        search_scrn.clear_widgets()
        
        
        df = read_excel(path)
        
        df['en_stock'] = df['en_stock'].fillna(0)
        
        df['vendu'] = df['en_stock']
        df['prix'] = df['en_stock']
        df['dernier_achat'] = df['en_stock']
        df['gpu'] = df['gpu'].fillna('STANDARD')
        df['dernier_achat'] = df['dernier_achat'].fillna('N/A')
        df['commentaire'] = df['commentaire'].fillna('N/A')
        df['batterie'] = df['batterie'].fillna('Bien')
        df['taux_de_change'] = df['taux_de_change'].fillna(1)
        df['charges'] = df['charges'].fillna(0)
        
        to_numeric(df['vendu'])
        to_numeric(df['taux_de_change'])
        to_numeric(df['en_stock'])
        to_numeric(df['charges'])
        
        charge = df['charges'].head(1)
        
        charge = float(charge)/df.shape[0]
        
        df['charges'] = charge
        
        df['prix_achat'] = df['prix_achat']*df['taux_de_change']+charge

        data = df.to_dict(orient="records")
        
        for record in data:
            record['Ref'] = self.gen_ref(record['marque'])
            record['prix'] = 0
            record['vendu'] = 0
            record['dernier_achat'] = ''

        self.products.insert_many(data)
        
        products = self.get_products()
        usertable  = DataTable(table=products)
        content.add_widget(usertable)
        
        product_table2  = DataTable(table=products)
        order_scrn.add_widget(product_table2)
        
        product_table3 = DataTable(table=products)
        supplier_scrn.add_widget(product_table3)
        
        product_table4 = DataTable(table=products)
        search_scrn.add_widget(product_table4)
        
        self.success_popup(f"{len(data)} Produits Importés")
        
        return 0
    
    
    def product_exist(self, ref):
        i = 0
        product = self.products.find({'Ref': f'{ref}'})
        try:
            product = product[0]
            return True 
        except IndexError:
            return False 


    def user_exist(self, username):
        users = self.users
        i = 0
        for p in users.find():
            if p['user_name'] == f"{username}":
                i += 1
        if i == 0:
            return False
        else:
            return True
    
    
    def get_users(self):
        users = self.users
        
        _users = {}
        _users['first_names'] = {}
        _users['last_names'] = {}
        _users['user_names'] = {}
        _users['passwords'] = {}
        _users['designations'] = {}
        
        first_names = []
        last_names = []
        user_names = []
        passwords = []
        designations = []

        for user in users.find():
            first_names.append(user['first_name'])
            last_names.append(user['last_name'])
            user_names.append(user['user_name'])
            pwd = user['password']
            if len(pwd) > 10:
                pwd  = f"{pwd[:10]}..."
            passwords.append(pwd)
            designations.append(user['designation'])

        for c, v in enumerate(first_names):
            _users['first_names'][c] = first_names[c]
            _users['last_names'][c] = last_names[c]
            _users['user_names'][c] = user_names[c]
            _users['passwords'][c] = passwords[c]
            _users['designations'][c] = designations[c]
        return _users

    
    def get_products(self, for_export=False):
        products = self.products
        
        if for_export:
            _stocks = {}
            _stocks['Ref'] = {}
            _stocks['marque'] = {}
            _stocks['modele'] = {}
            _stocks['cpu'] = {}
            _stocks['ram'] = {}
            _stocks['stockage'] = {}
            _stocks['gpu'] = {}
            _stocks['batterie'] = {}
            _stocks["cout d'achat"] = {}
            _stocks['prix de vente'] = {}
            _stocks['gain'] = {}
            _stocks['en_stock'] = {}
            _stocks['vendu'] = {}
            _stocks['commande'] = {}
            _stocks['fournisseur'] = {}
            _stocks['dernier_achat'] = {}
            _stocks['commentaire'] = {}

            Ref = []
            prix = []
            prix_achat = []
            marque = []
            modele = []
            cpu = []
            ram = []
            gpu = []
            stockage = []
            batterie = []
            en_stock = []
            vendu = []
            commande = []
            fournisseur = []
            dernier_achat = []
            commentaire = []
            
            gain = []

            for product in products.find():
                Ref.append(product['Ref'])
                prix.append(product['prix'])
                try:
                    prix_achat.append(product['prix_achat'])
                except KeyError:
                    prix_achat.append('')
                marque.append(product['marque'])
                modele.append(product['modele'])
                cpu.append(product['cpu'])
                ram.append(product['ram'])
                gpu.append(product['gpu'])
                stockage.append(product['stockage'])
                batterie.append(product['batterie'])
                try:    
                    en_stock.append(product['en_stock'])
                except KeyError:
                    en_stock.append('')

                try:    
                    vendu.append(product['vendu'])
                except KeyError:
                    vendu.append('')

                commande.append(product['commande'])

                try:    
                    fournisseur.append(product['fournisseur'])
                except KeyError:
                    fournisseur.append('')

                try:    
                    dernier_achat.append(product['dernier_achat'])
                except KeyError:
                    dernier_achat.append('')

                try:    
                    commentaire.append(product['commentaire'])
                except KeyError:
                    commentaire.append('****')
                    
                    
                try:
                    tbuy_p = float(product['prix_achat'])
                except:
                    tbuy_p = 0
                    
                try:
                    tsell_p = float(product['prix'])
                except:
                    tsell_p = 0
                
                gain.append(tsell_p-tbuy_p)

            sum_purchase_price = []
            sum_sold_list = []
            sold_items = []
            in_stock_items = []
            in_stock_items_price = []

            for c, v in enumerate(Ref):
                _stocks['Ref'][c] = Ref[c]
                _stocks['marque'][c] = marque[c]
                _stocks['modele'][c] = modele[c]
                _stocks['cpu'][c] = cpu[c]
                _stocks['ram'][c] = ram[c]
                _stocks['stockage'][c] = stockage[c]
                _stocks['gpu'][c] = gpu[c]
                _stocks['batterie'][c] = batterie[c]
                _stocks['prix de vente'][c] = prix[c]
                _stocks["cout d'achat"][c] = prix_achat[c]
                _stocks["gain"][c] = gain[c]
                _stocks['en_stock'][c] = en_stock[c]
                _stocks['vendu'][c] = vendu[c]
                _stocks['commande'][c] = commande[c]
                _stocks['dernier_achat'][c] = dernier_achat[c]
                _stocks['commentaire'][c] = commentaire[c]
                _stocks['fournisseur'][c] = fournisseur[c]
            
                if vendu[c] == '':
                    vendu[c] = 0

                if en_stock[c] == '':
                    en_stock[c] = 0
        else:
            stats = self.ids.scrn_product_stats
            stats.clear_widgets()
            _stocks = {}
            _stocks['Ref'] = {}
            _stocks['designation'] = {}
            _stocks["cout d'achat"] = {}
            _stocks['prix de vente'] = {}
            _stocks['gain'] = {}
            _stocks['en_stock'] = {}
            _stocks['vendu'] = {}
            _stocks['commande'] = {}
            _stocks['fournisseur'] = {}
            _stocks['dernier_achat'] = {}
            _stocks['commentaire'] = {}

            Ref = []
            prix = []
            prix_achat = []
            marque = []
            modele = []
            cpu = []
            ram = []
            gpu = []
            stockage = []
            batterie = []
            en_stock = []
            vendu = []
            commande = []
            fournisseur = []
            dernier_achat = []
            commentaire = []
            gain = []

            for product in products.find():
                Ref.append(product['Ref'])
                prix.append(product['prix'])
                try:
                    prix_achat.append(product['prix_achat'])
                except KeyError:
                    prix_achat.append('')
                marque.append(product['marque'])
                modele.append(product['modele'])
                cpu.append(product['cpu'])
                ram.append(product['ram'])
                gpu.append(product['gpu'])
                stockage.append(product['stockage'])
                batterie.append(product['batterie'])
                try:    
                    en_stock.append(product['en_stock'])
                except KeyError:
                    en_stock.append('')

                try:    
                    vendu.append(product['vendu'])
                except KeyError:
                    vendu.append('')

                commande.append(product['commande'])

                try:    
                    fournisseur.append(product['fournisseur'])
                except KeyError:
                    fournisseur.append('')

                try:    
                    dernier_achat.append(product['dernier_achat'])
                except KeyError:
                    dernier_achat.append('')

                try:    
                    commentaire.append(product['commentaire'])
                except KeyError:
                    commentaire.append('****')
                
                try:
                    tbuy_p = float(product['prix_achat'])
                except:
                    tbuy_p = 0
                    
                try:
                    tsell_p = float(product['prix'])
                except:
                    tsell_p = 0
                
                gain.append(tsell_p-tbuy_p)

            sum_purchase_price = []
            sum_sold_list = []
            sold_items = []
            in_stock_items = []
            in_stock_items_price = []

            for c, v in enumerate(Ref):
                _stocks['Ref'][c] = Ref[c]
                _stocks['designation'][c] = f"{marque[c]} {modele[c]} | {cpu[c]} | {ram[c]}GB\n{stockage[c]} | {gpu[c]} | {batterie[c]}"
                _stocks['prix de vente'][c] = prix[c]
                _stocks["cout d'achat"][c] = prix_achat[c]
                _stocks["gain"][c] = gain[c]
                _stocks['en_stock'][c] = en_stock[c]
                _stocks['vendu'][c] = vendu[c]
                _stocks['commande'][c] = commande[c]
                _stocks['dernier_achat'][c] = dernier_achat[c]
                _stocks['commentaire'][c] = commentaire[c]
                _stocks['fournisseur'][c] = fournisseur[c]
            
                if vendu[c] == '':
                    vendu[c] = 0

                if en_stock[c] == '':
                    en_stock[c] = 0

                if int(vendu[c]) >= 1:
                    sum_sold_list.append(int(vendu[c])*float(prix[c]))
                    sold_items.append(int(vendu[c]))
                    sum_purchase_price.append(int(vendu[c])*float(prix_achat[c]))

                if int(en_stock[c]) >= 1:
                    in_stock_items_price.append(int(en_stock[c])*float(prix_achat[c]))
                    in_stock_items.append(int(en_stock[c]))
                    sum_purchase_price.append(int(en_stock[c])*float(prix_achat[c]))

            # key numbers
            total_order_price = sum(sum_purchase_price)
            in_stock_items = sum(in_stock_items)
            in_stock_items_price = sum(in_stock_items_price)
            sold_items = sum(sold_items)
            sold_products_price = float(sum(sum_sold_list))
            profit = sold_products_price-total_order_price

            totalorderpricelabel = Label(text=f'Prix Totale:\n{total_order_price} DH', bold=True, color=(0,0,0,1))
            instockitemslabel = Label(text=f'Produits en stock:\n{in_stock_items}', bold=True, color=(0,0,0,1))
            instockitemspricelabel = Label(text=f'Prix des produits en stock:\n{in_stock_items_price} DH', bold=True, color=(0,0,0,1))
            solditemslabel = Label(text=f'Produits Vendu:\n{sold_items}', bold=True, color=(0,0,0,1))
            soldproductspricelabel = Label(text=f'Prix des produits Vendu:\n{sold_products_price} DH', bold=True, color=(0,0,0,1))
            profitlabel = Label(text=f'Profit:\n{profit} DH', bold=True, color=(0,0,0,1))

            stats.add_widget(totalorderpricelabel)
            stats.add_widget(instockitemslabel)
            stats.add_widget(instockitemspricelabel)
            stats.add_widget(solditemslabel)
            stats.add_widget(soldproductspricelabel)
            stats.add_widget(profitlabel)
   
        return _stocks
    
    
    def export_excel_fields(self):
        target = self.ids.ops_fields_p
        target.clear_widgets()
        

        crud_export_instock =  Button(text='Exporter Les Produits En Stock', size_hint_x=1/8, width=100,
                                    on_release=lambda x: self.export_xlsx(True),background_color=(0.184, 0.216, 0.231), background_normal='')
        
        crud_export_all =  Button(text='Exporter Tous Les Produits', size_hint_x=1/8, width=100,
                                on_release=lambda x: self.export_xlsx(),background_color=(0.184, 0.216, 0.231), background_normal='')
        
        crud_close_p =  Button(text='Fermer', size_hint_x=1/8, width=100,
                            on_release=lambda x: self.ids.ops_fields_p.clear_widgets(),
                            background_color=(0.184, 0.216, 0.231), background_normal='')
        
        target.add_widget(crud_export_instock)
        target.add_widget(crud_export_all)
        target.add_widget(crud_close_p)
        
        return 0

    
    def add_product_fields(self):
        target = self.ids.ops_fields_p
        target.clear_widgets()
        
        crud_marque = TextInput(hint_text='Marque', multiline=False)
        crud_modele = TextInput(hint_text='Modèle', multiline=False)
        crud_cpu = TextInput(hint_text='CPU', multiline=False)
        crud_ram = TextInput(hint_text='RAM', multiline=False)
        crud_gpu = TextInput(hint_text='GPU', multiline=False)
        crud_stockage = TextInput(hint_text='stockage', multiline=False)
        crud_batterie = TextInput(hint_text='Batterie', multiline=False)
        crud_price = TextInput(hint_text='Prix', multiline=False)
        crud_buy_price = TextInput(hint_text="cout d'achat", multiline=False)
        exchange_rate = TextInput(hint_text="Taux de change", multiline=False)
        fees = TextInput(hint_text="Charges", multiline=False)
        crud_stock = TextInput(hint_text='En stock', multiline=False)
        crud_sold = TextInput(hint_text='Vendu', multiline=False)
        crud_order = TextInput(hint_text='Commande', multiline=False)
        crud_fournisseur = TextInput(hint_text='Fournisseur', multiline=False)
        crud_last_purchase = TextInput(hint_text='Dernier Achat', multiline=False)
        crud_comment = TextInput(hint_text='Commentaire', multiline=False)

        crud_submit_p =  Button(text='Ajouter Produit', size_hint_x=1/8, width=100,
                                on_release=lambda x:self.add_product(self.gen_ref(crud_marque.text), crud_marque.text,
                                                                    crud_modele.text, crud_cpu.text, crud_ram.text, crud_gpu.text,
                                                                    crud_stockage.text, crud_batterie.text, crud_price.text,
                                                                    crud_buy_price.text, exchange_rate.text, fees.text, crud_stock.text, crud_sold.text, crud_order.text, 
                                                                    crud_fournisseur.text, crud_last_purchase.text, crud_comment.text), background_color=(0.184, 0.216, 0.231), background_normal='')
        
        
        target.add_widget(crud_marque)
        target.add_widget(crud_modele)
        target.add_widget(crud_cpu)
        target.add_widget(crud_ram)
        target.add_widget(crud_gpu)
        target.add_widget(crud_stockage)
        target.add_widget(crud_batterie)
        target.add_widget(crud_price)
        target.add_widget(crud_buy_price)
        target.add_widget(exchange_rate)
        target.add_widget(fees)
        target.add_widget(crud_stock)
        target.add_widget(crud_sold)
        target.add_widget(crud_order)
        target.add_widget(crud_fournisseur)
        target.add_widget(crud_last_purchase)
        target.add_widget(crud_comment)
        target.add_widget(crud_submit_p)        
        return 0


    def add_product(self, ref, marque, modele, cpu, ram, gpu, stockage, batterie, price, buy_price, exchange, fee, stock, sold, order, supplier, last_purchase, comment):
        if ref == '':
            self.missing_field_popup(field='Réf')
            return 0
        if self.product_exist(ref):
            self.error_popup('La Référence existe déjà')
            return 0
        if marque == '':
            self.missing_field_popup(field='marque')
            return 0
        if modele == '':
            self.missing_field_popup(field='Modèle')
            return 0
        if cpu == '':
            self.missing_field_popup(field='CPU')
            return 0
        if ram == '':
            self.missing_field_popup(field='RAM')
            return 0
        if gpu == '':
            gpu == 'STANDARD'
        if stockage == '':
            self.missing_field_popup(field='Stockage')
            return 0
        if batterie == '':
            self.missing_field_popup(field='Batterie')
            return 0
        
        if stock == '':
            self.missing_field_popup(field='En Stock')
            return 0
        if order == '':
            self.missing_field_popup(field='Commande')
            return 0
        
        if price == '':
            price = '0'
        if exchange == '':
            exchange = 1
        
        if fee == '':
            fee = 0
        
        exchange = float(exchange)
        buy_price = float(buy_price)*exchange+float(fee)
        buy_price = "{:.2f}".format(buy_price)
            
        if supplier == '':
            supplier = '?'
        
        order_scrn = self.ids.scrn_order_contents
        order_scrn.clear_widgets()
        
        content = self.ids.scrn_product_contents
        content.clear_widgets()
        
        supplier_scrn = self.ids.scrn_supplier_contents
        supplier_scrn.clear_widgets()
        
        self.products.insert_one({'Ref':ref, 'marque': marque,
                            'modele':modele, 'cpu': cpu, 'ram':ram, 'gpu':gpu, 'stockage': stockage, 'batterie':batterie, 'prix': price,
                            'prix_achat': buy_price, 'en_stock': stock, 'vendu': sold, 'commande': order, 'dernier_achat': last_purchase,
                            'commentaire': comment, 'fournisseur': supplier, 'charges': fee})
        
        products = self.get_products()
        usertable  = DataTable(table=products)
        content.add_widget(usertable)
        
        product_table2  = DataTable(table=products)
        order_scrn.add_widget(product_table2)
        
        
        product_table3 = DataTable(table=products)
        supplier_scrn.add_widget(product_table3)
        
        search_scrn = self.ids.scrn_search_contents
        search_scrn.clear_widgets()
        
        product_table4 = DataTable(table=products)
        search_scrn.add_widget(product_table4)
        
        self.success_popup("Produit Ajouté")
        
        return 0
    

    def u_p_autofill(self, code):
        if code == '':
            return 0
        else:
            product = self.get_product(order=code)
            if len(product['Ref']) <= 0:
                self.error_popup("Réference Invalide")
                return 0
            else:
                self.crud_marque.text = str(product['marque'])
                self.crud_modele.text = str(product['modele'])
                self.crud_cpu.text = str(product['cpu'])
                self.crud_ram.text = str(product['ram'])
                self.crud_gpu.text = str(product['gpu'])
                self.crud_stockage.text = str(product['stockage'])
                self.crud_batterie.text = str(product['batterie'])
                self.crud_price.text = str(product['prix de vente'])
                self.crud_buy_price.text = str(float(product["cout d'achat"])-float(product["charges"]))
                self.fee.text = str(product["charges"])
                self.crud_stock.text = str(product['en_stock'])
                self.crud_sold.text = str(product['vendu'])
                self.crud_order.text = str(product['commande'])
                self.crud_fournisseur.text = str(product['fournisseur'])
                self.crud_comment.text = str(product['commentaire'])
                return 0


    def update_product_fields(self):
        target = self.ids.ops_fields_p
        target.clear_widgets()
        
        crud_code = TextInput(hint_text='Réf', multiline=False)
        self.crud_marque = TextInput(hint_text='Marque', multiline=False)
        self.crud_modele = TextInput(hint_text='Modèle', multiline=False)
        self.crud_cpu = TextInput(hint_text='CPU', multiline=False)
        self.crud_ram = TextInput(hint_text='RAM', multiline=False)
        self.crud_gpu = TextInput(hint_text='GPU', multiline=False)
        self.crud_stockage = TextInput(hint_text='stockage', multiline=False)
        self.crud_batterie = TextInput(hint_text='Batterie', multiline=False)
        self.crud_price = TextInput(hint_text='Prix', multiline=False)
        self.crud_buy_price = TextInput(hint_text="cout d'achat", multiline=False)
        self.crud_exchange_rate = TextInput(hint_text="Taux de change", multiline=False)
        self.fee = TextInput(hint_text="Charges", multiline=False)
        self.crud_stock = TextInput(hint_text='En stock', multiline=False)
        self.crud_sold = TextInput(hint_text='Vendu', multiline=False)
        self.crud_order = TextInput(hint_text='Commande', multiline=False)
        self.crud_fournisseur = TextInput(hint_text='Fournisseur', multiline=False)
        self.crud_comment = TextInput(hint_text='Commentaire', multiline=False)
        crud_search =  Button(text='Recherche', on_release=lambda x:self.u_p_autofill(crud_code.text),
                              background_color=(0.184, 0.216, 0.231), background_normal='')
        crud_submit_p =  Button(text='Modifier Produit', size_hint_x=1/8, width=100,# price, buy_price, exchange, stock, sold, order, supplier, last_purchase, comment)
                                on_release=lambda x:self.update_product(crud_code.text, self.crud_marque.text, self.crud_modele.text, self.crud_cpu.text,
                                                                        self.crud_ram.text, self.crud_gpu.text, self.crud_stockage.text, self.crud_batterie.text,
                                                                        self.crud_price.text, self.crud_buy_price.text, self.crud_exchange_rate.text, self.fee.text, self.crud_stock.text, self.crud_sold.text,
                                                                        self.crud_order.text, self.crud_fournisseur.text, "OMO", self.crud_comment.text), 
                                background_color=(0.184, 0.216, 0.231), background_normal='')
        
        target.add_widget(crud_code)
        target.add_widget(crud_search)
        target.add_widget(self.crud_marque)
        target.add_widget(self.crud_modele)
        target.add_widget(self.crud_cpu)
        target.add_widget(self.crud_ram)
        target.add_widget(self.crud_gpu)
        target.add_widget(self.crud_stockage)
        target.add_widget(self.crud_batterie)
        target.add_widget(self.crud_price)
        target.add_widget(self.crud_buy_price)
        target.add_widget(self.crud_exchange_rate)
        target.add_widget(self.fee)
        target.add_widget(self.crud_stock)
        target.add_widget(self.crud_sold)
        target.add_widget(self.crud_order)
        target.add_widget(self.crud_fournisseur)
        target.add_widget(self.crud_comment)
        target.add_widget(crud_submit_p)
        crud_code.focus = True
        return 0

  
    def update_product(self, ref, marque, modele, cpu, ram, gpu, stockage, batterie, price, buy_price, exchange, fee, stock, sold, order, supplier, last_purchase, comment):
        if ref == '':
            self.missing_field_popup(field="Réf")
            return 0
        
        if not self.product_exist(ref):
            self.error_popup("la référence n' existe pas")
            return 0
        
        content = self.ids.scrn_product_contents
        content.clear_widgets()
        
        order_scrn = self.ids.scrn_order_contents
        order_scrn.clear_widgets()
        
        
        supplier_scrn = self.ids.scrn_supplier_contents
        supplier_scrn.clear_widgets()
        
        if price == '':
            price = '0'
        
        if exchange == '':
            exchange = 1
        
        if fee == '':
            fee = 0
        
        exchange = float(exchange)
        buy_price = float(buy_price)*exchange+float(fee)
        buy_price = "{:.2f}".format(buy_price)
        
        if supplier == '':
            supplier = '?'
        
        if comment == '':
            comment = ' '    
        
        
        if last_purchase == "OMO":
            self.products.update_one({'Ref':ref}, {'$set':{'marque': marque, 'modele':modele, 'cpu': cpu, 'ram':ram, 'gpu':gpu, 'stockage': stockage, 'batterie':batterie, 'prix': price,'prix_achat': buy_price, 'en_stock': stock, 'vendu': sold, 'commande': order, 'fournisseur': supplier, 'commentaire':comment, 'charges': fee}})
        else:
            self.products.update_one({'Ref':ref}, {'$set':{'marque': marque, 'modele':modele, 'cpu': cpu, 'ram':ram, 'gpu':gpu, 'stockage': stockage, 'batterie':batterie, 'prix': price,'prix_achat': buy_price, 'en_stock': stock, 'vendu': sold, 'commande': order, 'fournisseur': supplier, 'dernier_achat': last_purchase, 'commentaire':comment, 'charges': fee}})
        
        products = self.get_products()
        usertable  = DataTable(table=products)
        content.add_widget(usertable)
        
        product_table2  = DataTable(table=products)
        order_scrn.add_widget(product_table2)
        
        product_table3 = DataTable(table=products)
        supplier_scrn.add_widget(product_table3)
        
        
        search_scrn = self.ids.scrn_search_contents
        search_scrn.clear_widgets()
        
        product_table4 = DataTable(table=products)
        search_scrn.add_widget(product_table4)

        self.success_popup("Produit Modifié")
        
        return 0

    def remove_product_fields(self):
        target = self.ids.ops_fields_p
        target.clear_widgets()
        crud_ref = TextInput(hint_text="Référence", multiline=False, width=100, height=10)
        crud_submit =  Button(text='Supprimer Produit', size_hint_x=None, width=150, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.remove_product(crud_ref.text))
        crud_close =  Button(text='Fermer', size_hint_x=None, width=100, on_release=lambda x: self.ids.ops_fields_p.clear_widgets(),
                            background_color=(0.184, 0.216, 0.231), background_normal='')
        spacer = Label(text='', size_hint_x=.6)
        target.add_widget(crud_ref)
        target.add_widget(crud_submit)
        target.add_widget(crud_close)
        target.add_widget(spacer)
        
        return 0


    def return_product_fields(self):
        target = self.ids.ops_fields_p
        target.clear_widgets()
        crud_ref = TextInput(hint_text="Référence", multiline=False, width=50, height=5)
        crud_qty = TextInput(hint_text="Quantité", multiline=False, width=50, height=5)
        crud_submit =  Button(text='Retourner Produit', size_hint_x=None, width=150, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.return_product(crud_ref.text, crud_qty.text))
        crud_close =  Button(text='Fermer', size_hint_x=None, width=100, on_release=lambda x: self.ids.ops_fields_p.clear_widgets(),
                            background_color=(0.184, 0.216, 0.231), background_normal='')
        spacer = Label(text='', size_hint_x=.6)
        target.add_widget(crud_ref)
        target.add_widget(crud_qty)
        target.add_widget(crud_submit)
        target.add_widget(crud_close)
        target.add_widget(spacer)
        return 0
    
    
    def return_product(self, ref, minus):
        if ref == '' or minus == '':
            return 0
        
        if not self.product_exist(ref):
            self.error_popup("la référence n' existe pas")
            return 0
        
        content = self.ids.scrn_product_contents
        content.clear_widgets()
        
        order_scrn = self.ids.scrn_order_contents
        order_scrn.clear_widgets()
        
        
        supplier_scrn = self.ids.scrn_supplier_contents
        supplier_scrn.clear_widgets()
        
        prdct = self.products.find({'Ref': f'{ref}'})
        try:
            prdct = prdct[0]
        except IndexError:
            self.error_popup("Réference Invalid")
            return 0
        
        if prdct['vendu'] == '':
            prdct['vendu'] = 0
        if prdct['en_stock'] == '':
            prdct['en_stock'] = 0
        
        new_en_stock = int(prdct['en_stock']) + int(minus)
        sold = int(prdct['vendu']) - int(minus)
        self.products.update_one({'Ref':ref}, {'$set':{'en_stock': new_en_stock, 'vendu': sold, 'prix': '0'}})
        
        products = self.get_products()
        usertable  = DataTable(table=products)
        content.add_widget(usertable)
        
        product_table2  = DataTable(table=products)
        order_scrn.add_widget(product_table2)
        
        product_table3 = DataTable(table=products)
        supplier_scrn.add_widget(product_table3)
        
        search_scrn = self.ids.scrn_search_contents
        search_scrn.clear_widgets()
        
        product_table4 = DataTable(table=products)
        search_scrn.add_widget(product_table4)
        
        return 0


    def remove_product(self, ref):
        if ref == '':
            self.missing_field_popup(field="Nom d'utilisateur")
            return 0
        if not self.product_exist(ref):
            self.error_popup("Le Produit n' existe pas")
            return 0
        
        content = self.ids.scrn_product_contents
        content.clear_widgets()
        
        order_scrn = self.ids.scrn_order_contents
        order_scrn.clear_widgets()
        
        supplier_scrn = self.ids.scrn_supplier_contents
        supplier_scrn.clear_widgets()
        
        self.products.delete_many({"Ref":ref})
        
        products = self.get_products()
        usertable  = DataTable(table=products)
        content.add_widget(usertable)
        
        product_table2  = DataTable(table=products)
        order_scrn.add_widget(product_table2)
        
        search_scrn = self.ids.scrn_search_contents
        search_scrn.clear_widgets()
        
        product_table4 = DataTable(table=products)
        search_scrn.add_widget(product_table4)
        
        product_table3 = DataTable(table=products)
        supplier_scrn.add_widget(product_table3)
            
        self.success_popup(f"le Produit {ref} a été Supprimé")
        return 0
    
    
    def add_user_fields(self):
        target = self.ids.ops_fields
        target.clear_widgets()
        crud_first = TextInput(hint_text='Prénom', multiline=False)
        crud_last = TextInput(hint_text='Nom', multiline=False)
        crud_user = TextInput(hint_text="Nom d'utilisateur", size_hint_x=3/2, multiline=False)
        crud_pwd = TextInput(hint_text='Mot de passe', multiline=False, password=True)
        crud_des = Spinner(text='Operateur', values=['Operateur', 'Administrateur'],
                        background_color=(0.184, 0.216, 0.231), background_normal='')
        crud_submit =  Button(text='Ajouter', size_hint_x=None, width=100,
                            background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.add_user(crud_first.text, crud_last.text, crud_user.text, crud_pwd.text, crud_des.text))
        crud_close =  Button(text='Fermer', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.ids.ops_fields.clear_widgets())
        target.add_widget(crud_first)
        target.add_widget(crud_last)
        target.add_widget(crud_user)
        target.add_widget(crud_pwd)
        target.add_widget(crud_des)
        target.add_widget(crud_submit)
        target.add_widget(crud_close)
        return 0


    def add_charge_fields(self):
        target = self.ids.charges_ops_fields
        target.clear_widgets()
        motif = TextInput(hint_text='Motif', size_hint_x=3/2, multiline=False)
        montant = TextInput(hint_text='Montant', multiline=False)
        date = TextInput(hint_text="Date", multiline=False)
        date.text = str(datetime.now().strftime('%d-%m-%Y'))
        crud_submit =  Button(text='Ajouter', size_hint_x=None, width=100,
                            background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.add_charge(self.gen_charge_ref(), motif.text, montant.text, date.text))
        crud_close =  Button(text='Fermer', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.ids.charges_ops_fields.clear_widgets())
        target.add_widget(motif)
        target.add_widget(montant)
        target.add_widget(date)
        target.add_widget(crud_submit)
        target.add_widget(crud_close)
        
        motif.focus = True
        
        return 0


    def update_charge_fields(self):
        target = self.ids.charges_ops_fields
        target.clear_widgets()
        ref = TextInput(hint_text='Référence', size_hint_x=3/2, multiline=False)
        motif = TextInput(hint_text='Motif', size_hint_x=3/2, multiline=False)
        montant = TextInput(hint_text='Montant', multiline=False)
        date = TextInput(hint_text="Date", multiline=False)
        crud_submit =  Button(text='Modifier', size_hint_x=None, width=100,
                            background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.update_charge(ref.text, motif.text, montant.text, date.text))
        crud_close =  Button(text='Fermer', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.ids.charges_ops_fields.clear_widgets())
        target.add_widget(ref)
        target.add_widget(motif)
        target.add_widget(montant)
        target.add_widget(date)
        target.add_widget(crud_submit)
        target.add_widget(crud_close)
        ref.focus = True
        return 0


    def remove_charge_fields(self):
        target = self.ids.charges_ops_fields
        target.clear_widgets()
        crud_user = TextInput(hint_text="Référence", multiline=False)
        crud_submit =  Button(text='Supprimer', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.remove_charge(crud_user.text))
        crud_close =  Button(text='Fermer', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.ids.charges_ops_fields.clear_widgets())
        target.add_widget(crud_user)
        target.add_widget(crud_submit)
        target.add_widget(crud_close)
        
        crud_user.focus = True
        return 0
    

    def remove_charge(self, ref):
        if ref == '':
            self.missing_field_popup(field="Référence")
            return 0
        if not self.charge_exist(ref):
            self.error_popup("Le référence n' existe pas")
            return 0
        content = self.ids.scrn_charges_contents
        content.clear_widgets()
        self.charges.delete_many({"Ref":ref})
        charges = self.get_charges()
        
        chargetable  = DataTable(table=charges)
        content.add_widget(chargetable)
        
        self.success_popup(f"La charge {ref} a été Supprimé")
        return 0


    def transaction_exist(self, ref):
        tr = self.transactions.find({'Ref': f'{ref}'})
        try:
            tr = tr[0]
            return True 
        except IndexError:
            return False
        
    def open_reciept(self):
        ref = self.ids.tr_id_inp.text
        if self.transaction_exist(ref):
            try:
                os.startfile(f".\\recu\\{ref}.txt")
                return 0
            except FileNotFoundError:
                self.error_popup("Fichier Introuvable")
                return 0
        else:
            self.error_popup("Réference Invalide")
            return 0

    
    def open_invoice(self):
        ref = self.ids.tr_id_inp.text
        if self.transaction_exist(ref):
            try:
                os.startfile(f".\\factures\\{ref}.pdf")
                return 0
            except FileNotFoundError:
                self.error_popup("Fichier Introuvable")
                return 0
        else:
            self.error_popup("Réference Invalide")
            return 0


    def update_charge(self, ref, motif, montant, date):
        if ref == '':
            self.missing_field_popup(field="Nom d'utilisateur")
            return 0
        if not self.charge_exist(ref):
            self.error_popup("Réference Erroné")
            return 0
        content = self.ids.scrn_charges_contents
        content.clear_widgets()
        self.charges.update_one({'Ref':ref}, {'$set':{'motif':motif, 'montant': montant, 'date':date}})
        charges = self.get_charges()
        chargetable  = DataTable(table=charges)
        content.add_widget(chargetable)

        self.success_popup(f"La charge {ref} Modifié")
        return 0


    def charge_exist(self, ref):
        i = 0
        product = self.charges.find({'Ref': f'{ref}'})
        try:
            product = product[0]
            return True 
        except IndexError:
            return False 


    def search_charges(self, month, year):
        
        charges_scrn = self.ids.scrn_charges_contents
        charges_scrn.clear_widgets()
        self.ids.scrn_charges_stats.clear_widgets()
        
        charges = self.charges

        if month != '':
            reg = f'{month}-{year}$'
        else:
            reg = f'{year}$'

        _charges = {}
        _charges['Ref'] = {}
        _charges['montant'] = {}
        _charges['motif'] = {}
        _charges['date'] = {}

        ref = []
        motifs = []
        montants = []
        dates = []
        for user in charges.find():
            if re.search(reg, user['date']) != None:
                ref.append(user['Ref'])
                motifs.append(user['motif'])
                montants.append(user['montant'])
                dates.append(user['date'])

        for c, v in enumerate(ref):
            _charges['Ref'][c] = ref[c]
            _charges['motif'][c] = motifs[c]
            _charges['montant'][c] = montants[c]
            _charges['date'][c] = dates[c]

        somme_charges = []
        
        for element in montants:
            somme_charges.append(float(element))
            
        somme_charges = sum(somme_charges)
        chargeslabel = Label(text=f'Somme: {somme_charges} DH', bold=True, color=(0,0,0,1))
        self.ids.scrn_charges_stats.add_widget(chargeslabel)
        
        prod_table  = DataTable(table=_charges)
        charges_scrn.add_widget(prod_table)
        
        return _charges


    def get_charges(self):
        
        charges = self.charges
        
        self.ids.scrn_charges_stats.clear_widgets()

        _charges = {}
        _charges['Ref'] = {}
        _charges['montant'] = {}
        _charges['motif'] = {}
        _charges['date'] = {}

        ref = []
        motifs = []
        montants = []
        dates = []
        for user in charges.find():
            ref.append(user['Ref'])
            motifs.append(user['motif'])
            montants.append(user['montant'])
            dates.append(user['date'])

        for c, v in enumerate(ref):
            _charges['Ref'][c] = ref[c]
            _charges['motif'][c] = motifs[c]
            _charges['montant'][c] = montants[c]
            _charges['date'][c] = dates[c]
            
        somme_charges = []
        
        for element in montants:
            somme_charges.append(float(element))
            
        somme_charges = sum(somme_charges)
        chargeslabel = Label(text=f'Somme: {somme_charges} DH', bold=True, color=(0,0,0,1))
        self.ids.scrn_charges_stats.add_widget(chargeslabel)
        
        return _charges
    

    def get_transactions(self):
        charges = self.transactions
        self.ids.scrn_transaction_stats.clear_widgets()
        
        _charges = {}
        _charges['Ref'] = {}
        _charges['Totale_HT'] = {}
        _charges['TVA'] = {}
        _charges['Totale_TTC'] = {}
        _charges['date'] = {}
        
        ref = []
        total_ht = []
        tva = []
        total_ttc = []
        date = []

        for user in charges.find():
            ref.append(user['Ref'])
            total_ht.append(user['Totale_HT'])
            tva.append(user['TVA'])
            total_ttc.append(user['Totale_TTC'])
            date.append(user['date'])

        for c, v in enumerate(ref):
            _charges['Ref'][c] = ref[c]
            _charges['Totale_HT'][c] = total_ht[c]
            _charges['TVA'][c] = tva[c]
            _charges['Totale_TTC'][c] = total_ttc[c]
            _charges['date'][c] = date[c]
            
        somme_ttc = []
        somme_tva = []
        somme_ht = []
        
        for c, element in enumerate(total_ht):
            somme_ttc.append(float(total_ttc[c]))
            somme_tva.append(float(tva[c]))
            somme_ht.append(float(total_ht[c]))
        
        somme_ttc = sum(somme_ttc)
        somme_tva = sum(somme_tva)
        somme_ht = sum(somme_ht)
        
        htlabel = Label(text=f'Somme HT: {somme_ht} DH', bold=True, color=(0,0,0,1))
        ttclabel = Label(text=f'Somme TTC: {somme_ttc} DH', bold=True, color=(0,0,0,1))
        tvalabel = Label(text=f'Somme TVA: {somme_tva} DH', bold=True, color=(0,0,0,1))
        
        self.ids.scrn_transaction_stats.add_widget(htlabel)
        self.ids.scrn_transaction_stats.add_widget(tvalabel)
        self.ids.scrn_transaction_stats.add_widget(ttclabel)
        
        return _charges


    def get_clients(self):
        charges = self.clients
                
        _charges = {}
        _charges['Ref'] = {}
        _charges['RS'] = {}
        _charges['RC'] = {}
        _charges['IF'] = {}
        _charges['ICE'] = {}
        _charges['date'] = {}
        
        ref = []
        rs = []
        rc = []
        idf = []
        ice = []
        date = []

        for user in charges.find():
            ref.append(user['Ref'])
            rs.append(user['RS'])
            rc.append(user['RC'])
            idf.append(user['IF'])
            ice.append(user['ICE'])
            date.append(user['date'])

        for c, v in enumerate(ref):
            _charges['Ref'][c] = ref[c]
            _charges['RS'][c] = rs[c]
            _charges['RC'][c] = rc[c]
            _charges['IF'][c] = idf[c]
            _charges['ICE'][c] = ice[c]
            _charges['date'][c] = date[c]
        
        return _charges
    
   
    def add_charge(self, ref, motif, montant, date):
        if ref == '':
            return 0

        if motif == '':
            self.missing_field_popup(field='Motif')
            return 0
        if montant == '':
            self.missing_field_popup(field='Modèle')
            return 0
        if date == '':
            self.missing_field_popup(field='CPU')
            return 0
        
        charge_scrn = self.ids.scrn_charges_contents
        charge_scrn.clear_widgets()
        
        self.charges.insert_one({'Ref':ref, 'motif': motif, 'montant':montant, 'date': date})
        
        charges = self.get_charges()

        
        charges_table  = DataTable(table=charges)
        charge_scrn.add_widget(charges_table)
        
        self.success_popup("Charge Ajouté")
        
        return 0
    
    
    def add_user(self, first, last, user,  pwd, des):
        if user == '':
            self.missing_field_popup(field="Nom d'utilisateur")
            return 0
        if self.user_exist(user):
            self.error_popup('Utilisateur existe déjà')
            return 0
        if pwd == '':
            self.missing_field_popup(field="Mot de Passe")
            return 0
        content = self.ids.scrn_contents
        content.clear_widgets()
        pwd = hashlib.sha256(pwd.encode()).hexdigest()
        self.users.insert_one({'first_name':first, 'last_name': last,
                            'user_name':user, 'password': pwd, 'designation':des, 'date':datetime.now()})
        users = self.get_users()
        usertable  = DataTable(table=users)
        content.add_widget(usertable)
        
        self.success_popup(f"Utilisateur {user} Ajouté")
        
        return  0


    def update_user_fields(self):
        target = self.ids.ops_fields
        target.clear_widgets()
        crud_first = TextInput(hint_text='Prénom', multiline=False)
        crud_last = TextInput(hint_text='Nom', multiline=False)
        crud_user = TextInput(hint_text="Nom d'utilisateur", size_hint_x=3/2, multiline=False)
        crud_pwd = TextInput(hint_text='Mot de passe', multiline=False, password=True)
        crud_des = Spinner(text='Operateur', values=['Operateur', 'Administrateur'], background_color=(0.184, 0.216, 0.231), background_normal='')
        crud_submit =  Button(text='Modifier', size_hint_x=None, width=100,
                            background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.update_user(crud_first.text, crud_last.text, crud_user.text, crud_pwd.text, crud_des.text))
        crud_close =  Button(text='Fermer', background_color=(0.184, 0.216, 0.231), background_normal='',
                            size_hint_x=None, width=100, on_release=lambda x: self.ids.ops_fields.clear_widgets())
        target.add_widget(crud_user)
        target.add_widget(crud_first)
        target.add_widget(crud_last)
        target.add_widget(crud_pwd)
        target.add_widget(crud_des)
        target.add_widget(crud_submit)
        target.add_widget(crud_close)

        return 0

    
    def update_user(self, first, last, user,  pwd, des):
        if user == '':
            self.missing_field_popup(field="Nom d'utilisateur")
            return 0
        if not self.user_exist(user):
            self.error_popup("Utilisateur n' existe pas")
            return 0
        content = self.ids.scrn_contents
        content.clear_widgets()
        pwd = hashlib.sha256(pwd.encode()).hexdigest()
        self.users.update_one({'user_name':user}, {'$set':{'first_name':first, 'last_name': last,
                            'user_name':user, 'password': pwd, 'designation':des, 'date':datetime.now()}})
        users = self.get_users()
        usertable  = DataTable(table=users)
        content.add_widget(usertable)
        
        self.success_popup(f"Utilisateur {user} Modifié")
        
        return  0
    
    
    def remove_user_fields(self):
        target = self.ids.ops_fields
        target.clear_widgets()
        crud_user = TextInput(hint_text="Nom d'utilisateur", multiline=False)
        crud_submit =  Button(text='Supprimer', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.remove_user(crud_user.text))
        crud_close =  Button(text='Fermer', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.ids.ops_fields.clear_widgets())
        target.add_widget(crud_user)
        target.add_widget(crud_submit)
        target.add_widget(crud_close)
        
        return 0
    
    
    def search_charge_fields(self):
        target = self.ids.charges_ops_fields
        target.clear_widgets()
        crud_month = TextInput(hint_text="Mois", multiline=False)
        crud_year = TextInput(hint_text="Année", multiline=False)
        crud_submit =  Button(text='Rechercher', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.search_charges(crud_month.text, crud_year.text))
        crud_reset =  Button(text='Effacer', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.charges_defaults())
        crud_close =  Button(text='Fermer', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.ids.charges_ops_fields.clear_widgets())
        target.add_widget(crud_month)
        target.add_widget(crud_year)
        target.add_widget(crud_submit)
        target.add_widget(crud_reset)
        target.add_widget(crud_close)

        return 0
    
    
    def search_transaction_fields(self):
        
        target = self.ids.transaction_ops_fields
        target.clear_widgets()
        crud_month = TextInput(hint_text="Mois", multiline=False)
        crud_year = TextInput(hint_text="Année", multiline=False)
        crud_submit =  Button(text='Rechercher', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.search_transactions(crud_month.text, crud_year.text))
        crud_reset =  Button(text='Effacer', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.transaction_defaults())
        crud_close =  Button(text='Fermer', size_hint_x=None, width=100, background_color=(0.184, 0.216, 0.231), background_normal='',
                            on_release=lambda x: self.ids.transaction_ops_fields.clear_widgets())
        target.add_widget(crud_month)
        target.add_widget(crud_year)
        target.add_widget(crud_submit)
        target.add_widget(crud_reset)
        target.add_widget(crud_close)

        return 0
    

    def transaction_defaults(self):
        # scrn_transaction_contents
        tr_scrn = self.ids.scrn_transaction_contents
        tr_scrn.clear_widgets()
        products = self.get_transactions()
        prod_table  = DataTable(table=products)
        tr_scrn.add_widget(prod_table)
        
        return 0
    
    def charges_defaults(self):
        charges_scrn = self.ids.scrn_charges_contents
        charges_scrn.clear_widgets()
        products = self.get_charges()
        prod_table  = DataTable(table=products)
        charges_scrn.add_widget(prod_table)
        
        return 0


    def search_transactions(self, month, year):
        tr_scrn = self.ids.scrn_transaction_contents
        tr_scrn.clear_widgets()
        self.ids.scrn_transaction_stats.clear_widgets()
        
        charges = self.transactions
        
        if month != '':
            reg = f'{month}-{year}$'
        else:
            reg = f'{year}$'
        
        _charges = {}
        _charges['Ref'] = {}
        _charges['Totale_HT'] = {}
        _charges['TVA'] = {}
        _charges['Totale_TTC'] = {}
        _charges['date'] = {}
        
        ref = []
        total_ht = []
        tva = []
        total_ttc = []
        date = []
        

        for user in charges.find():
            if re.search(reg, user['date']) != None:
                ref.append(user['Ref'])
                total_ht.append(user['Totale_HT'])
                tva.append(user['TVA'])
                total_ttc.append(user['Totale_TTC'])
                date.append(user['date'])

        for c, v in enumerate(ref):
            _charges['Ref'][c] = ref[c]
            _charges['Totale_HT'][c] = total_ht[c]
            _charges['TVA'][c] = tva[c]
            _charges['Totale_TTC'][c] = total_ttc[c]
            _charges['date'][c] = date[c]
            
        somme_ttc = []
        somme_tva = []
        somme_ht = []
        
        for c, element in enumerate(total_ht):
            somme_ttc.append(float(total_ttc[c]))
            somme_tva.append(float(tva[c]))
            somme_ht.append(float(total_ht[c]))
        
        somme_ttc = sum(somme_ttc)
        somme_tva = sum(somme_tva)
        somme_ht = sum(somme_ht)
        
        htlabel = Label(text=f'Somme HT: {somme_ht} DH', bold=True, color=(0,0,0,1))
        ttclabel = Label(text=f'Somme TTC: {somme_ttc} DH', bold=True, color=(0,0,0,1))
        tvalabel = Label(text=f'Somme TVA: {somme_tva} DH', bold=True, color=(0,0,0,1))
        
        self.ids.scrn_transaction_stats.add_widget(htlabel)
        self.ids.scrn_transaction_stats.add_widget(tvalabel)
        self.ids.scrn_transaction_stats.add_widget(ttclabel)
        
        # scrn_transaction_contents
        prod_table  = DataTable(_charges)
        tr_scrn.add_widget(prod_table)
        
        return _charges

    
    def remove_user(self, username):
        if username == '':
            self.missing_field_popup(field="Nom d'utilisateur")
            return 0
        if not self.user_exist(username):
            self.error_popup("L' utilisateur n' existe pas")
            return 0
        content = self.ids.scrn_contents
        content.clear_widgets()
        self.users.delete_many({"user_name":username})
        users = self.get_users()
        usertable  = DataTable(table=users)
        content.add_widget(usertable)
        
        self.success_popup(f"Utilisateur {username} Supprimé")
        return 0
    
    
    def change_screen(self, instance):
        if instance.text == 'Utilisateurs':
            self.ids.scrn_mngr.current = 'scrn_content'
        elif instance.text == 'Inventaire':
            self.ids.scrn_mngr.current = 'scrn_product_content'
        elif instance.text == 'Commandes':
            self.ids.scrn_mngr.current = 'scrn_order_content'
        elif instance.text == 'Calculateur':
            self.ids.scrn_mngr.current = 'scrn_calc_content'
        elif instance.text == 'Charges':
            self.ids.scrn_mngr.current = 'scrn_charges_content' 
        elif instance.text == 'Recherche':
            self.ids.scrn_mngr.current = 'scrn_search_content'
        elif instance.text == 'Fournisseurs':
            self.ids.scrn_mngr.current = 'scrn_supplier_content'
        elif instance.text == 'Transactions':
            self.ids.scrn_mngr.current = 'scrn_transaction_content'
        elif instance.text == 'Clients':
            self.ids.scrn_mngr.current = 'scrn_clients_content'
        return 0


class AdminApp(App):
    def build(self):
        return AdminWindow()


if __name__=="__main__":
    aa = AdminApp()
    aa.run()
    